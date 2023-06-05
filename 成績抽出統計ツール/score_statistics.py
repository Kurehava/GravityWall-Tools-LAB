from os import walk, system, listdir
from os.path import splitext, isfile, isdir, split
from random import sample
from string import ascii_letters, digits
from time import strftime, localtime

import openpyxl
import numpy as np
from openpyxl.styles.alignment import Alignment
from statistics import stdev

ERROR_P = '[\033[091mERROR\033[0m]'
WARN__P = '[\033[093mWARN-\033[0m]'
INFO__P = '[\033[092mINFO-\033[0m]'


class LibraryPathSplit:
    def __init__(self, file_path):
        random_str = ''.join(sample(ascii_letters + digits, 2))
        now_time = strftime("%Y%m%d%H%M%S", localtime())
        nr = f'{now_time}_{random_str}'

        self.file_path = f"{file_path}"
        self.nr = nr

    def root(self):
        return split(str(self.file_path))[0]

    def ext(self):
        return splitext(split(self.file_path)[1])[1].split('.')[-1]

    def file_name_ext(self):
        return split(self.file_path)[1]

    def file_name(self):
        return splitext(split(self.file_path)[1])[0]

    def file_name_mix(self):
        return f"{splitext(split(self.file_path)[1])[0]}_{self.nr}"

    def root_file_name(self):
        return f'{split(self.file_path)[0]}/{splitext(split(self.file_path)[1])[0]}'

    def root_file_name_mix(self):
        return f'{split(self.file_path)[0]}/{splitext(split(self.file_path)[1])[0]}_{self.nr}'

    def all(self):
        temp = LibraryPathSplit(self.file_path)
        return [self.file_path,
                temp.root(),
                temp.ext(),
                temp.file_name_ext(),
                temp.file_name(),
                temp.file_name_mix(),
                temp.root_file_name(),
                temp.root_file_name_mix()]


# 获取有效数据的行与列下标
def SEARCH_DATA(sd_str, EXCEL_MAX_ROW, EXCEL_SHEET_DATAS):
    type_flag = 0
    result_row_start = 0
    result_row_end = 0
    end_str = ""

    if str(type(sd_str))[8:-2] != "str":
        (sd_str, end_str) = sd_str
        type_flag = 1
    row_content_chk_flag = 1
    for v_row in range(1, EXCEL_MAX_ROW):
        row_content = EXCEL_SHEET_DATAS.cell(row=v_row, column=1).value
        if str(row_content).find(sd_str) != -1:
            result_row_start = v_row + 1
            row_content_chk_flag = 2
        if row_content_chk_flag == 2:
            if type_flag == 0:
                if str(row_content) == "None":
                    result_row_end = v_row
                    break
            elif type_flag == 1:
                if str(row_content) == end_str:
                    result_row_end = v_row
                    break
    try:
        return [result_row_start, result_row_end]
    except:
        raise KeyError("%s %s Not found element." % (sd_str, end_str))


def multiple_file_split(input_path):
    """
    多文件终端拖入路径分离函数

    :return: file_list, fail_list
    """

    file_list = []
    fail_list = []
    addr_start = -1
    addr_end = 0
    addr_flags = 0

    input_path = f'{input_path}/'
    for length in range(len(input_path)):
        if input_path[length] == '/' and addr_flags != 1:
            # print(f'DEBUG-flag-1:{addr_flags}::{length}->{addr_start}-{addr_end}')
            addr_flags = 1
            addr_start = length
        if input_path[length] == '.' and addr_flags == 1:
            # print(f'DEBUG-flag-2:{addr_flags}::{length}->{addr_start}-{addr_end}')
            addr_flags = 2
        if input_path[length:length + 2] == ' /' and addr_flags == 2:
            # print(f'DEBUG-flag-3:{addr_flags}::{length}->{addr_start}-{addr_end}')
            addr_flags = 3
            addr_end = length
        elif input_path[length:length + 2] == ' /' and addr_flags != 2:
            # print(f'DEBUG-flag-4:{addr_flags}::{length}->{addr_start}-{addr_end}')
            addr_start = 0
            addr_end = 0
            addr_flags = 0
        if addr_start != -1 and addr_end != 0 and addr_flags == 3:
            # print(f'DEBUG-flag-5:{addr_flags}::{length}->{addr_start}-{addr_end}')
            file_path = input_path[addr_start:addr_end]
            file_path = file_path.replace("\\", "")
            # print(f'DEBUG-file_path:{file_path}')
            if isfile(file_path) and len(file_path) > 10:
                file_list.append(file_path)
            else:
                fail_list.append(file_path)
            addr_start = 0
            addr_end = 0
            addr_flags = 0
    if not file_list:
        print(f'{ERROR_P}: multiple_file_split \n'
              f'{ERROR_P}: 没有检测到拖入文件.\n')
        input("按下回车返回主菜单")
        return False
    return file_list, fail_list


def folder_file_statistic(INPUT_PATH):
    file_list = []
    fail_list = []

    for fn in listdir(INPUT_PATH):
        file_path = f"{INPUT_PATH}/{fn}"
        if isfile(file_path) and splitext(file_path)[1].upper() == ".XLSX":
            file_list.append(file_path)
        else:
            fail_list.append(file_path)
    if len(file_list) == 0:
        print(f'{ERROR_P}: folder_file_statistic \n'
              f'{ERROR_P}: 文件夹内可用文件为空.\n')
        raise FileNotFoundError("Not found .xlsx file.")
    return file_list, fail_list


if __name__ == "__main__":
    # 定义文件路径与表名称
    EXCEL_PATH = input("Input CSV FILES PATH or Put CSV FILES or CSV FILES FOLDER in THIS WINDOWS.\n>>")
    temp_EP = EXCEL_PATH.strip(' ').strip('"').strip('\'')
    if isdir(temp_EP):
        EXCEL_PATH = temp_EP
        excel_file_list, fail_path_list = folder_file_statistic(EXCEL_PATH)
        for paths in excel_file_list:
            file_name = LibraryPathSplit(paths).file_name()
            if file_name == '.DS_Store' or file_name[0] == '~' or (file_name[0:2] != "No" and file_name[-11:] != "-TESTS.xlsx"):
                excel_file_list.pop(excel_file_list.index(paths))
    else:
        excel_file_list, fail_path_list = multiple_file_split(EXCEL_PATH)

    SAVE_PATH = f"{LibraryPathSplit(excel_file_list[0]).root()}/result.xlsx"

    title_list = ['氏名', '学籍番号']

    stu_name_ = []
    stu_score = []
    stu_numb_ = []

    add_count = 0
    len_max_cell = 0

    excel_file_list.sort()
    for file_path in excel_file_list:
        add_count += 1
        Lps = LibraryPathSplit(file_path)
        test_number = str(Lps.file_name()).replace('-TESTS', '')
        title_list.append(test_number)
        meta_data = openpyxl.load_workbook(file_path)
        sheet_data = meta_data.worksheets[0]

        max_row = sheet_data.max_row
        max_col = sheet_data.max_column

        for index in range(2, max_row + 1):
            name_ = sheet_data.cell(column=1, row=index).value
            stdno = int(sheet_data.cell(column=2, row=index).value)
            score = float(sheet_data.cell(column=5, row=index).value)

            if name_ not in stu_name_:
                stu_name_.append(name_)
                stu_numb_.append(stdno)
                if add_count == 1:
                    stu_score.append([score])
                else:
                    temp = []
                    for _ in range(1, add_count):
                        temp.append(0.0)
                    temp.append(score)
                    stu_score.append(temp)
            else:
                target_index = stu_name_.index(name_)
                stu_score[target_index].append(score)

        for list_index in range(len(stu_score)):
            list_elem = stu_score[list_index]
            if len(list_elem) != add_count:
                stu_score[list_index].append(0.0)

    # title_list.append('合計値')
    title_list.extend(['合計値', '偏差値'])
    elem_sum_list = []
    for index, elem in enumerate(stu_score):
        elem_sum = sum(elem)
        stu_score[index].append(elem_sum)
        elem_sum_list.append(elem_sum)

    avg = sum(elem_sum_list) / len(stu_score)
    std = stdev(elem_sum_list)

    title_list.append('')
    elem_std_list = []
    for index, elem in enumerate(elem_sum_list):
        temp = (elem - avg) / std * 10 + 50
        stu_score[index].append(temp)
        elem_std_list.append(temp)

    temp = [np.mean([stu_score[index][idx] for index in range(len(stu_score)) if stu_score[index][idx] != 0]) for idx in range(len(stu_score[0]))]
    stu_score.append(temp)
    temp = ['' for _ in range(len(stu_score[0]) - 3)]
    # stu_score.append(temp + ['平均値', avg, sum(elem_std_list) / len(elem_std_list)])
    # stu_score.append(temp + ['標準偏差', std])
    stu_name_.extend([''])
    stu_numb_.extend(['平均値(0除く)'])

    write_list = [title_list]
    for index in range(len(stu_name_)):
        temp = [stu_name_[index], stu_numb_[index]]
        temp.extend(stu_score[index])
        write_list.append(temp)

    # 写文件
    NEW_WBK = openpyxl.Workbook()
    NEW_SHT = NEW_WBK.active
    NEW_SHT.title = "result"
    LEN = len(write_list)
    EXIT_FLAG = 0

    for data in write_list:
        NEW_SHT.append(data)

    NEW_SHT.column_dimensions['A'].width = 12
    NEW_SHT.column_dimensions['B'].width = 12

    for index in range(1, len(write_list) + 1):
        NEW_SHT[f"B{index}"].alignment = Alignment(horizontal='left')

    NEW_WBK.save(SAVE_PATH)

    # system("osascript -e 'tell application \"Terminal\" to close first window' & exit")
