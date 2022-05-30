###################################################
# POWER BY GravityWallToolsDevelopmentLAB Project #
###################################################
###############################################################
# "Information Security" Achievement Statistics Tool          #
# Powered by Kureha Belonging to KanagawaUniversity MoritaLab #
# Github:https://github.com/Kurehava                          #
###############################################################
import csv
import os
import sys

import numpy as np
import openpyxl
import pandas as pd
import xlwt

os.system('clear')

SEARCH_STR = ["[成績一覧]", "[ユーザ毎の回答リスト]"]
COLS_RANGE = [[2, 3, 4, 7, 8], 5]
DATA = []

TESTS_NUMB = input("How many tests is this?\n>>").strip("'")
EXCEL_PATH = input("Input CSV FILE PATH or Put CSV FILE in THIS WINDOWS.\n>>").strip("'").strip(' ')
EXCEL_FILE_NAME = str(os.path.split(EXCEL_PATH)).split(".")[0]
EXCEL_SHEET = 1
EXCEL_XLS_NAME = EXCEL_PATH.replace(EXCEL_PATH.split(".")[-1], "") + "xls"
EXCEL_XLSX_NAME = EXCEL_PATH.replace(EXCEL_PATH.split(".")[-1], "") + "xlsx"
NEW_XLSX_PATH = EXCEL_PATH.replace(EXCEL_PATH.split("/")[-1], "") + "No.%s-TESTS.xlsx" % TESTS_NUMB

if os.path.splitext(EXCEL_PATH)[1] in {".csv", ".CSV"}:
    print("PROCESSING...")
    with open(EXCEL_PATH, 'r', encoding='shift-jis', errors="ignore") as f:
        read = csv.reader(f)
        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet('data')
        line = 0
        for lines in read:
            r = 0
            for i in lines:
                sheet.write(line, r, i)
                r = r + 1
            line = line + 1
        workbook.save(EXCEL_XLS_NAME)
    data = pd.DataFrame(pd.read_excel(EXCEL_XLS_NAME, engine="xlrd"))
    data.to_excel(EXCEL_XLSX_NAME, index=False)
    EXCEL_PATH = EXCEL_PATH.replace(EXCEL_PATH.split(".")[-1], "xlsx")
    os.remove(EXCEL_XLS_NAME)
else:
    print("FILE IS NOT A CSV FILE.\n")

EXCEL_DATA = openpyxl.load_workbook(EXCEL_PATH)
EXCEL_SHEET_DATA = EXCEL_DATA.worksheets[0]

EXCEL_MAX_ROW = EXCEL_SHEET_DATA.max_row
EXCEL_MAX_COL = EXCEL_SHEET_DATA.max_column

def SEARCH_DATA(sd_str):
    type_flag = 0
    result_row_start = 0
    result_row_end = 0
    end_str = ""

    if str(type(sd_str))[8:-2] != "str":
        (sd_str, end_str) = sd_str
        type_flag = 1
    row_content_chk_flag = 1
    for v_row in range(1, EXCEL_MAX_ROW):
        row_content = EXCEL_SHEET_DATA.cell(row=v_row, column=1).value
        if str(row_content).find(sd_str) != -1:
            result_row_start = v_row + 1
            row_content_chk_flag = 2
        if row_content_chk_flag == 2:
            if type_flag == 0:
                if str(row_content) == "None":
                    result_row_end = v_row
                    break
            elif type_flag == 1:
                if str(row_content) == end_str:
                    result_row_end = v_row
                    break
    try:
        return [result_row_start, result_row_end]
    except:
        raise KeyError("%s %s Not found element." % (sd_str, end_str))

DATA_ARR_COUNT = -1
COLS_ARR_COUNT = -1
ARR_END = 0
for STR_NAME in SEARCH_STR:
    (start, end) = SEARCH_DATA(STR_NAME)
    COLS_ARR_COUNT += 1
    if str(type(COLS_RANGE[COLS_ARR_COUNT]))[8:-2] == "int":
        ARR_START = COLS_RANGE[COLS_ARR_COUNT]
        for vir_row in range(start, end):
            for vir_col in range(ARR_START, EXCEL_MAX_COL):
                if str(EXCEL_SHEET_DATA.cell(row=start, column=vir_col).value) == 'None' and vir_col > ARR_END:
                    ARR_END = vir_col
        COLS_RANGE.pop(COLS_ARR_COUNT)
        COLS_RANGE.append([])
        for element in range(ARR_START, ARR_END):
            COLS_RANGE[COLS_ARR_COUNT].append(element)
    elif str(type(COLS_RANGE[COLS_ARR_COUNT]))[8:-2] != "list":
        raise TypeError("The elements in the array COLS_RANGE[] must be 'list' or 'int'.")
    for col_num in COLS_RANGE[COLS_ARR_COUNT]:
        DATA.append([])
        DATA_ARR_COUNT += 1
        for row in range(start, end):
            DATA[DATA_ARR_COUNT].append(EXCEL_SHEET_DATA.cell(row=row, column=col_num).value)

DATA_NAME_NP = np.array(DATA[0])
GET_REPEAT_NAME = []
GET_REPEAT_INDEX = []

for DATA_NAME_COUNT in range(len(DATA[0])):
    if DATA[0].count(DATA[0][DATA_NAME_COUNT]) != 1 and GET_REPEAT_NAME.count(DATA[0][DATA_NAME_COUNT]) == 0:
        GET_REPEAT_NAME.append(DATA[0][DATA_NAME_COUNT])

for GRN in GET_REPEAT_NAME:
    GET_REPEAT_LETTER = np.where(DATA_NAME_NP == GRN)
    GRN_ARR = []
    for elem in np.flipud(GET_REPEAT_LETTER[0]):
        GRN_ARR.append(elem)
    for elem_GRN in range(len(GRN_ARR)):
        if not DATA[5][GRN_ARR[elem_GRN]] == "未解答":
            GRN_ARR.pop(elem_GRN)
            break
        if elem_GRN == len(GRN_ARR) and DATA[5][GRN_ARR[elem_GRN]] == "未解答":
            GRN_ARR.pop(0)
            break
    for trans in GRN_ARR:
        GET_REPEAT_INDEX.append(trans)
    GET_REPEAT_INDEX.sort()

for DATA_SUB_COUNT in range(len(DATA)):
    DATA_DEL_KEY_COUNT = 0
    for L_RAN in range(len(GET_REPEAT_INDEX)):
        DATA[DATA_SUB_COUNT].pop(GET_REPEAT_INDEX[L_RAN] - DATA_DEL_KEY_COUNT)
        DATA_DEL_KEY_COUNT += 1

NEW_WBK = openpyxl.Workbook()
NEW_SHT = NEW_WBK.active
DATA_ROW_WRITE = []
NEW_SHT.title = "第%s回" % TESTS_NUMB
LEN = len(DATA[0])
EXIT_FLAG = 0

for num in range(len(DATA)):
    if LEN != len(DATA[num]):
        print("Warning! Array subscripts are not aligned!")
        sys.exit(1)

DATA_MAX_LEN = 0
for DATA_MAX_LEN_CHK in range(len(DATA)):
    if len(DATA[DATA_MAX_LEN_CHK]) > DATA_MAX_LEN:
        DATA_MAX_LEN = len(DATA[DATA_MAX_LEN_CHK])

for element_count in range(DATA_MAX_LEN):
    for count in range(len(DATA)):
        try:
            DATA_ROW_WRITE.append(DATA[count][element_count])
        except:
            pass
    NEW_SHT.append(DATA_ROW_WRITE)
    DATA_ROW_WRITE = []

NEW_WBK.save(NEW_XLSX_PATH)
os.remove(EXCEL_XLSX_NAME)
os.system("osascript -e 'tell application \"Terminal\" to close first window' & exit")
