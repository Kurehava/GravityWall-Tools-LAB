###################################################
# POWER BY GravityWallToolsDevelopmentLAB Project #
###################################################
###############################################################
# "Information Security" Achievement Statistics System        #
# Powered by Kureha Belonging to KanagawaUniversity MoritaLab #
# Github:https://github.com/Kurehava                          #
###############################################################
import csv
import os
import sys

import numpy as np
import openpyxl
import pandas as pd
import xlwt
from statistics import stdev

os.system('clear')


# 按顺序
SEARCH_STR = ["[成績一覧]", "[ユーザ毎の回答リスト]"]
COLS_RANGE = [[2, 3, 4, 7, 8], 5]
# 血的教训 绝对不能用这种方式预定义数组 否则他们地址会一致 然后就会导致append一个数组变成append全数组
# DATA = [[]] * sum([len(COLS_RANGE[n]) for n in range(len(COLS_RANGE))])
DATA = []

# 定义文件路径与表名称
# EXCEL_PATH = "/Users/Kureha/Documents/Projects/TA-excel/test/answer-3.csv"
# EXCEL_PATH = "/Users/Kureha/Desktop/answer-1.csv"
TESTS_NUMB = input("How many tests is this?\n>>").strip("'")
EXCEL_PATH = input("Input CSV FILE PATH or Put CSV FILE in THIS WINDOWS.\n>>").strip("'").strip(' ')
EXCEL_FILE_NAME = str(os.path.split(EXCEL_PATH)).split(".")[0]
EXCEL_SHEET = 1
EXCEL_XLS_NAME = EXCEL_PATH.replace(EXCEL_PATH.split(".")[-1], "") + "xls"
EXCEL_XLSX_NAME = EXCEL_PATH.replace(EXCEL_PATH.split(".")[-1], "") + "xlsx"
NEW_XLSX_PATH = EXCEL_PATH.replace(EXCEL_PATH.split("/")[-1], "") + "No.%s-TESTS.xlsx" % TESTS_NUMB

# 文件类型转换 CSV->XLS->XLSX
if os.path.splitext(EXCEL_PATH)[1] in {".csv", ".CSV"}:
    print("PROCESSING...")
    with open(EXCEL_PATH, 'r', encoding='shift-jis', errors="ignore") as f:
        read = csv.reader(f)
        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet('data')
        line = 0
        for lines in read:
            r = 0
            for i in lines:
                sheet.write(line, r, i)
                r = r + 1
            line = line + 1
        workbook.save(EXCEL_XLS_NAME)
    data = pd.DataFrame(pd.read_excel(EXCEL_XLS_NAME, engine="xlrd"))
    data.to_excel(EXCEL_XLSX_NAME, index=False)
    EXCEL_PATH = EXCEL_PATH.replace(EXCEL_PATH.split(".")[-1], "xlsx")
    os.remove(EXCEL_XLS_NAME)
else:
    print("FILE IS NOT A CSV FILE.\n")

# 总体读取EXCEL指定表数据
EXCEL_DATA = openpyxl.load_workbook(EXCEL_PATH)
EXCEL_SHEET_DATA = EXCEL_DATA.worksheets[0]

# 获取表最大行与最大列
EXCEL_MAX_ROW = EXCEL_SHEET_DATA.max_row
EXCEL_MAX_COL = EXCEL_SHEET_DATA.max_column


# 获取有效数据的行与列下标
def SEARCH_DATA(sd_str):
    type_flag = 0
    result_row_start = 0
    result_row_end = 0
    end_str = ""

    if str(type(sd_str))[8:-2] != "str":
        (sd_str, end_str) = sd_str
        type_flag = 1
    row_content_chk_flag = 1
    for v_row in range(1, EXCEL_MAX_ROW):
        row_content = EXCEL_SHEET_DATA.cell(row=v_row, column=1).value
        if str(row_content).find(sd_str) != -1:
            result_row_start = v_row + 1
            row_content_chk_flag = 2
        if row_content_chk_flag == 2:
            if type_flag == 0:
                if str(row_content) == "None":
                    result_row_end = v_row
                    break
            elif type_flag == 1:
                if str(row_content) == end_str:
                    result_row_end = v_row
                    break
    try:
        return [result_row_start, result_row_end]
    except:
        raise KeyError("%s %s Not found element." % (sd_str, end_str))


def get_index(lists, key):
    gi_IDX = None
    gi_idx = None
    for idxs in range(len(lists)):
        try:
            gi_idx = lists[idxs].index(key)
            gi_IDX = idxs
            break
        except ValueError:
            continue
    if gi_IDX is None:
        raise ValueError(f"Not found {key}")
    return [gi_IDX, gi_idx]


# 获取有效数据
DATA_ARR_COUNT = -1
COLS_ARR_COUNT = -1
ARR_END = 0
for STR_NAME in SEARCH_STR:
    (start, end) = SEARCH_DATA(STR_NAME)
    COLS_ARR_COUNT += 1
    if str(type(COLS_RANGE[COLS_ARR_COUNT]))[8:-2] == "int":
        ARR_START = COLS_RANGE[COLS_ARR_COUNT]
        for vir_row in range(start, end):
            for vir_col in range(ARR_START, EXCEL_MAX_COL):
                if str(EXCEL_SHEET_DATA.cell(row=start, column=vir_col).value) == 'None' and vir_col > ARR_END:
                    ARR_END = vir_col
        COLS_RANGE.pop(COLS_ARR_COUNT)
        COLS_RANGE.append([])
        for element in range(ARR_START, ARR_END):
            COLS_RANGE[COLS_ARR_COUNT].append(element)
    elif str(type(COLS_RANGE[COLS_ARR_COUNT]))[8:-2] != "list":
        raise TypeError("The elements in the array COLS_RANGE[] must be 'list' or 'int'.")
    for col_num in COLS_RANGE[COLS_ARR_COUNT]:
        DATA.append([])
        DATA_ARR_COUNT += 1
        for row in range(start, end):
            DATA[DATA_ARR_COUNT].append(EXCEL_SHEET_DATA.cell(row=row, column=col_num).value)

# print(DATA)
# DATA(list)
# 0:氏名 1:学籍番号 2:解答日 3:得点率 4:偏差値 5-inf:解答

# 数组去重 选取得点率最高的数据
NAME = np.array(DATA[0])
REPEAT_NAME = []
save_index = []

# 抽出数组中有重复的名字
for INDEX in range(len(DATA[0])):
    if DATA[0].count(DATA[0][INDEX]) != 1 and REPEAT_NAME.count(DATA[0][INDEX]) == 0:
        REPEAT_NAME.append(DATA[0][INDEX])

# 获取需要去重的数据下标
for repeat_name in REPEAT_NAME:
    # 获取重复名字的下标数组(NP format) 重复名字包含第一个被存入数组的元素
    name_index = np.where(NAME == repeat_name)
    # 转换为list
    name_index_list = [elem for elem in np.flipud(name_index[0])]
    # 获取得点率
    scoring_rate = [DATA[3][elem] for elem in name_index_list]
    # 元数据为字符串，并且含有很多杂质，先去除杂质
    for INDEX in range(len(scoring_rate)):
        if scoring_rate[INDEX] == "未解答":
            scoring_rate[INDEX] = 0
        else:
            scoring_rate[INDEX] = int(scoring_rate[INDEX].replace(" ", "").replace("%", ""))
    # 返回最大值下标, 并存入save数组
    # 取得最大值，找到最大值的下标，这个下标对应名字列表的下标
    # 在名字列表里找到对应下标的DATA数组的下标，把DATA数组的下标存入数组
    name_index_list.pop(scoring_rate.index(max(scoring_rate)))
    save_index.extend(name_index_list)

# 清理数组
# 倒序获取DATA的下标，如果把DATA[index]转换为集合，要素数<=1，那么就跳过，反之存入数组
DATA = [DATA[index] for index in range(len(DATA) - 1, -1, -1) if len(set(DATA[index])) > 1]
DATA.reverse()

# 创建写入数组
save_index = [x for x in range(len(DATA[1])) if x not in save_index]
save_index.sort()
write_data = [[DATA[index - 1][INDEX] for index in range(1, len(DATA))] for INDEX in save_index]

# 计算
IDX_score = get_index(write_data, "<得点率>")[1]
score = [int(write_data[index][IDX_score].replace(" ", "").replace("%", "")) for index in range(len(write_data)) if index > 0]
AVG = np.mean(score)
STD = stdev(score)

IDX_std = get_index(write_data, "<偏差値>")[1]

for idx in range(len(score)):
    write_data[idx + 1][IDX_std] = (score[idx] - AVG) / STD * 10 + 50

# 写文件
NEW_WBK = openpyxl.Workbook()
NEW_SHT = NEW_WBK.active
DATA_ROW_WRITE = []
NEW_SHT.title = "第%s回" % TESTS_NUMB
LEN = len(DATA[0])
EXIT_FLAG = 0

for data in write_data:
    NEW_SHT.append(data)

NEW_WBK.save(NEW_XLSX_PATH)
os.remove(EXCEL_XLSX_NAME)
# os.system("osascript -e 'tell application \"Terminal\" to close first window' & exit")
sys.exit(0)
